from openpyxl import load_workbook
import pyautogui
from time import sleep
import pyperclip

arquivo = load_workbook('produtos_ficticios.xlsx')
aba_produtos = arquivo['Produtos']
ultima_linha = aba_produtos.max_row

for linha in range(2,ultima_linha+1):
    #pegar os itens e escrever nos input
    produto = aba_produtos[f'A{linha}'].value
    #produto
    pyperclip.copy(produto)
    pyautogui.click(1275,309,duration=1)
    sleep(1)
    pyautogui.hotkey('ctrl','v')

    #descrição
    descriçao = aba_produtos[f'B{linha}'].value
    pyperclip.copy(descriçao)
    pyautogui.click(1280,424,duration=1)
    pyautogui.hotkey('ctrl','v')

    #categoria
    categoria = aba_produtos[f"C{linha}"].value
    pyperclip.copy(categoria)
    pyautogui.click(1305,582,duration=1)
    pyautogui.hotkey('ctrl','v')

    #codigo do produto
    codigo_produto = aba_produtos[f'D{linha}'].value
    pyperclip.copy(codigo_produto)
    pyautogui.click(1309,692,duration=1)
    pyautogui.hotkey('ctrl','v')

    #peso
    peso = aba_produtos[f'E{linha}'].value
    pyperclip.copy(peso)
    pyautogui.click(1324,796,duration=1)
    pyautogui.hotkey('ctrl','v')

    #dimensões
    dimensoes = aba_produtos[f'F{linha}'].value
    pyperclip.copy(dimensoes)
    pyautogui.click(1301,903,duration=1)
    pyautogui.hotkey('ctrl','v')
    
    #clicar próximo
    pyautogui.click(1283,975,duration=1)
    sleep(2)
    
    #preço
    preco = aba_produtos[f'G{linha}'].value
    pyperclip.copy(preco)
    pyautogui.click(1280,340,duration=1)
    pyautogui.hotkey('ctrl','v')
    
    #quantidade do estoque
    quantidade_estoque = aba_produtos[f'H{linha}'].value
    pyperclip.copy(quantidade_estoque)
    pyautogui.click(1299,450,duration=1)
    pyautogui.hotkey('ctrl','v')

    #data de validade
    data_validade = aba_produtos[f'I{linha}'].value
    pyperclip.copy(data_validade)
    pyautogui.click(1296,553,duration=1)
    pyautogui.hotkey('ctrl','v')

    #cor
    cor = aba_produtos[f'J{linha}'].value
    pyperclip.copy(cor)
    pyautogui.click(1295,661,duration=1)
    pyautogui.hotkey('ctrl','v')
    
    #tamanho
    tamanho = aba_produtos[f'K{linha}'].value
    pyautogui.click(1311,765,duration=1)
    if tamanho == 'Pequeno':
        pyautogui.click(1300,810,duration=1)
    elif tamanho == 'Médio':
        pyautogui.click(1299,847)
    elif tamanho == 'Grande':
        pyautogui.click(1292,883)
    
    #material
    material = aba_produtos[f'L{linha}'].value
    pyperclip.copy(material)
    pyautogui.click(1284,873,duration=1)
    pyautogui.hotkey('ctrl','v')

    #clica em próximo
    pyautogui.click(1289,950,duration=1)

    #fabricante
    fabricante = aba_produtos[f'M{linha}'].value
    pyperclip.copy(fabricante)
    pyautogui.click(1292,360,duration=1)
    pyautogui.hotkey('ctrl','v')

    #pais de origem
    pais_origem = aba_produtos[f'N{linha}'].value
    pyperclip.copy(pais_origem)
    pyautogui.click(1291,470,duration=1)
    pyautogui.hotkey('ctrl','v')

    #observações
    observacoes = aba_produtos[f'O{linha}'].value
    pyperclip.copy(observacoes)
    pyautogui.click(1278,578,duration=1)
    pyautogui.hotkey('ctrl','v')

    #codigo de barra
    codigo_barra = aba_produtos[f'P{linha}'].value
    pyperclip.copy(codigo_barra)
    pyautogui.click(1283,743,duration=1)
    pyautogui.hotkey('ctrl','v')

    #localização do armazem
    localizacao_armazem = aba_produtos[f'Q{linha}'].value
    pyperclip.copy(localizacao_armazem)
    pyautogui.click(1286,851,duration=1)
    pyautogui.hotkey('ctrl','v')

    #clicar em concluir
    pyautogui.click(1299,928,duration=1)

    #clica em ok
    sleep(2)
    pyautogui.click(1781,255,duration=1)
    
    #cadastrar novo produto
    sleep(2)
    pyautogui.click(1560,643,duration=1)
    




    





    

